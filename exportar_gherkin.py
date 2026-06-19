#!/usr/bin/env python3
"""
Exporta escenarios Gherkin (.feature) a Excel.
Uso: python3 exportar_gherkin.py [directorio_gherkin] [salida.xlsx]
"""

import os
import re
import sys
from dataclasses import dataclass, field

import openpyxl
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

HEADERS = [
    "SCENARIO",
    "EXAMPLE",
    "MÓDULO",
    "FUNCIONALIDAD",
    "SUB FUNCIONALIDAD",
    "PATH",
    "FEATURE NAME",
    "BACKGROUND /\nPRECONDICIÓN",
    "SCENARIO NAME",
    "STEPS",
    "DATA DE PRUEBA",
    "HISTORIA/TAREA",
    "PLATAFORMA",
    "PRIORIDAD",
    "CRITICIDAD",
    "ESTADO DE EJECUCIÓN",
    "EJECUTADO POR",
    "FECHA DE EJECUCIÓN",
    "BUGS (LINK)",
]

STEP_KEYWORDS = (
    "Dado ", "Cuando ", "Entonces ", "Y ", "Pero ",
    "Given ", "When ", "Then ", "And ", "But ", "* ",
)

SCENARIO_PREFIXES = ("Escenario:", "Scenario:")
OUTLINE_PREFIXES = ("Esquema del escenario:", "Scenario Outline:", "Esquema de escenario:")
FEATURE_PREFIXES = ("Característica:", "Feature:")
BACKGROUND_PREFIXES = ("Antecedentes:", "Antecedente:", "Background:")
EXAMPLES_PREFIXES = ("Ejemplos:", "Examples:")


@dataclass
class Scenario:
    name: str
    tags: list = field(default_factory=list)
    steps: list = field(default_factory=list)
    is_outline: bool = False
    examples_headers: list = field(default_factory=list)
    examples_rows: list = field(default_factory=list)
    background_steps: list = field(default_factory=list)
    feature_name: str = ""
    filepath: str = ""


def _strip_prefix(text: str, prefixes: tuple) -> str:
    for p in prefixes:
        if text.startswith(p):
            return text[len(p):].strip()
    return text


def _parse_table_row(line: str) -> list:
    return [c.strip() for c in line.strip().strip("|").split("|")]


def parse_feature(filepath: str, rel_path: str = "") -> list[Scenario]:
    with open(filepath, encoding="utf-8") as f:
        lines = f.readlines()

    feature_name = ""
    background_steps: list[str] = []
    scenarios: list[Scenario] = []

    pending_tags: list[str] = []
    current: Scenario | None = None
    in_background = False
    in_examples = False
    examples_header_set = False

    for line in lines:
        stripped = line.strip()

        if not stripped or stripped.startswith("#"):
            continue

        # Tags
        if stripped.startswith("@"):
            pending_tags = re.findall(r"@[\w]+", stripped)
            continue

        # Feature
        if any(stripped.startswith(p) for p in FEATURE_PREFIXES):
            feature_name = _strip_prefix(stripped, FEATURE_PREFIXES)
            pending_tags = []
            continue

        # Background
        if any(stripped.startswith(p) for p in BACKGROUND_PREFIXES):
            in_background = True
            in_examples = False
            current = None
            pending_tags = []
            continue

        # Scenario
        if any(stripped.startswith(p) for p in SCENARIO_PREFIXES):
            if current:
                scenarios.append(current)
            current = Scenario(
                name=_strip_prefix(stripped, SCENARIO_PREFIXES),
                tags=pending_tags[:],
                feature_name=feature_name,
                filepath=rel_path,
            )
            pending_tags = []
            in_background = False
            in_examples = False
            continue

        # Scenario Outline
        if any(stripped.startswith(p) for p in OUTLINE_PREFIXES):
            if current:
                scenarios.append(current)
            current = Scenario(
                name=_strip_prefix(stripped, OUTLINE_PREFIXES),
                tags=pending_tags[:],
                is_outline=True,
                feature_name=feature_name,
                filepath=rel_path,
            )
            pending_tags = []
            in_background = False
            in_examples = False
            continue

        # Examples
        if any(stripped.startswith(p) for p in EXAMPLES_PREFIXES):
            in_examples = True
            examples_header_set = False
            continue

        # Table rows
        if stripped.startswith("|"):
            cols = _parse_table_row(stripped)
            if in_examples and current:
                if not examples_header_set:
                    current.examples_headers = cols
                    examples_header_set = True
                else:
                    current.examples_rows.append(cols)
            elif in_background:
                background_steps.append(stripped)
            elif current and not in_examples:
                # Data table attached to last step
                if current.steps:
                    current.steps[-1] += "\n      " + stripped
            continue

        # Steps
        if any(stripped.startswith(kw) for kw in STEP_KEYWORDS):
            if in_background:
                background_steps.append(stripped)
            elif current and not in_examples:
                current.steps.append(stripped)
            continue

        # Docstrings
        if stripped.startswith('"""') or stripped.startswith("'''"):
            continue

    if current:
        scenarios.append(current)

    for s in scenarios:
        s.background_steps = background_steps[:]

    return scenarios


def collect_rows(gherkin_dir: str) -> list[dict]:
    base = os.path.abspath(gherkin_dir)
    feature_files = sorted(
        os.path.join(root, f)
        for root, _, files in os.walk(gherkin_dir)
        for f in files
        if f.endswith(".feature")
    )

    rows = []
    scenario_counter = 0

    for filepath in feature_files:
        rel_path = os.path.relpath(filepath, os.path.dirname(base))
        for scenario in parse_feature(filepath, rel_path):
            scenario_counter += 1

            base_row = {
                "n": scenario_counter,
                "path": scenario.filepath,
                "feature": scenario.feature_name,
                "background": "\n".join(scenario.background_steps),
                "name": scenario.name,
                "steps": "\n".join(scenario.steps),
                "tags": scenario.tags,
            }

            if scenario.is_outline and scenario.examples_rows:
                for ex_i, ex_row in enumerate(scenario.examples_rows, start=1):
                    pairs = [
                        f"{h}: {v}"
                        for h, v in zip(scenario.examples_headers, ex_row)
                    ]
                    rows.append({**base_row, "ex": ex_i, "data": "\n".join(pairs)})
            else:
                rows.append({**base_row, "ex": "", "data": ""})

    return rows


def _extract_levels(path: str) -> tuple[str, str, str]:
    """Extrae Módulo, Funcionalidad y Sub Funcionalidad del path del feature.

    gherkin/                         → nivel 0 (raíz, se ignora)
    gherkin/<modulo>/                → nivel 1 → Módulo
    gherkin/<modulo>/<func>/         → nivel 2 → Funcionalidad
    gherkin/<modulo>/<func>/<sub>/   → nivel 3 → Sub Funcionalidad
    """
    parts = path.replace("\\", "/").split("/")
    folders = parts[1:-1]  # descarta 'gherkin' (índice 0) y el archivo (último)
    modulo         = folders[0] if len(folders) > 0 else ""
    funcionalidad  = folders[1] if len(folders) > 1 else ""
    sub_func       = folders[2] if len(folders) > 2 else ""
    return modulo, funcionalidad, sub_func



def _prioridad(tags: list[str]) -> str:
    mapping = {
        "@prioridadBaja":    "Baja",
        "@prioridadMedia":   "Media",
        "@prioridadAlta":    "Alta",
        "@prioridadExtrema": "Extrema",
    }
    for tag in tags:
        if tag in mapping:
            return mapping[tag]
    return ""


def _criticidad(tags: list[str]) -> str:
    mapping = {
        "@c_baja":   "Baja",
        "@c_Media":  "Media",
        "@c_alta":   "Alta",
        "@c_extrema": "Extrema",
    }
    for tag in tags:
        if tag in mapping:
            return mapping[tag]
    return ""


def write_excel(rows: list[dict], output_path: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Escenarios"

    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=10)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
    ws.row_dimensions[1].height = 42

    fill_even = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
    fill_odd = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    cell_font = Font(size=9, name="Consolas")

    centered_headers = {
        "SCENARIO", "EXAMPLE", "MÓDULO", "FUNCIONALIDAD", "SUB FUNCIONALIDAD",
        "PATH", "FEATURE NAME", "BACKGROUND /\nPRECONDICIÓN", "SCENARIO NAME",
        "PRIORIDAD", "CRITICIDAD", "ESTADO DE EJECUCIÓN", "EJECUTADO POR", "FECHA DE EJECUCIÓN",
    }
    centered_cols = {col for col, h in enumerate(HEADERS, 1) if h in centered_headers}
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_top    = Alignment(vertical="top", wrap_text=True)

    for i, row in enumerate(rows, start=2):
        fill = fill_even if i % 2 == 0 else fill_odd
        modulo, funcionalidad, sub_func = _extract_levels(row["path"])
        values = [
            row["n"],
            row["ex"],
            modulo,
            funcionalidad,
            sub_func,
            row["path"],
            row["feature"],
            row["background"],
            row["name"],
            row["steps"],
            row["data"],
            "",   # HISTORIA/TAREA
            "",   # PLATAFORMA
            _prioridad(row["tags"]),    # PRIORIDAD
            _criticidad(row["tags"]),   # CRITICIDAD
            "PENDING",   # ESTADO DE EJECUCIÓN (dropdown)
            "",   # EJECUTADO POR
            "",   # FECHA DE EJECUCIÓN
            "",   # BUGS (LINK)
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.fill = fill
            cell.font = cell_font
            cell.alignment = align_center if col in centered_cols else align_top

        # Row height based on longest cell content
        max_lines = max(
            (str(v).count("\n") + 1 for v in values if v),
            default=1,
        )
        ws.row_dimensions[i].height = max(15, min(max_lines * 14, 250))

    col_widths = [10, 10, 20, 25, 25, 40, 30, 35, 45, 50, 35, 20, 15, 15, 15, 20, 20, 20, 20]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.freeze_panes = "A2"

    # Dropdowns de validación de datos
    last_row = len(rows) + 1

    def _add_dropdown(formula: str, col_name: str, error_msg: str):
        col_letter = get_column_letter(HEADERS.index(col_name) + 1)
        dv = DataValidation(
            type="list",
            formula1=formula,
            allow_blank=True,
            showErrorMessage=True,
            errorTitle="Valor inválido",
            error=error_msg,
        )
        ws.add_data_validation(dv)
        dv.add(f"{col_letter}2:{col_letter}{last_row}")

    _add_dropdown('"Baja,Media,Alta,Extrema"',    "PRIORIDAD",          "Seleccione: Baja, Media, Alta o Extrema")
    _add_dropdown('"Baja,Media,Alta,Extrema"',    "CRITICIDAD",         "Seleccione: Baja, Media, Alta o Extrema")
    _add_dropdown('"PENDING,In progress,PASSED,FAILED,BLOCKED"', "ESTADO DE EJECUCIÓN", "Seleccione: PENDING, In progress, PASSED, FAILED o BLOCKED")

    # Formato condicional por color
    col_prioridad = get_column_letter(HEADERS.index("PRIORIDAD") + 1)
    col_estado    = get_column_letter(HEADERS.index("ESTADO DE EJECUCIÓN") + 1)

    def _cond_color(col: str, value: str, bg: str, fg: str):
        ws.conditional_formatting.add(
            f"{col}2:{col}{last_row}",
            FormulaRule(
                formula=[f'${col}2="{value}"'],
                fill=PatternFill(start_color=bg, end_color=bg, fill_type="solid"),
                font=Font(size=9, name="Consolas", color=fg, bold=True),
            ),
        )

    # PRIORIDAD
    _cond_color(col_prioridad, "Baja",    "C6EFCE", "276221")  # verde
    _cond_color(col_prioridad, "Media",   "FFEB9C", "9C6500")  # amarillo
    _cond_color(col_prioridad, "Alta",    "FFDCA0", "7F3F00")  # naranja
    _cond_color(col_prioridad, "Extrema", "FFC7CE", "9C0006")  # rojo

    # ESTADO DE EJECUCIÓN
    _cond_color(col_estado, "PENDING",     "BFBFBF", "262626")  # gris
    _cond_color(col_estado, "In progress", "BDD7EE", "1F4E79")  # azul
    _cond_color(col_estado, "PASSED",      "C6EFCE", "276221")  # verde
    _cond_color(col_estado, "FAILED",      "FFC7CE", "9C0006")  # rojo
    _cond_color(col_estado, "BLOCKED",     "F4CCCC", "7F0000")  # rojo oscuro

    wb.save(output_path)
    print(f"Exportado: {output_path}  ({len(rows)} filas)")


def main():
    gherkin_dir = sys.argv[1] if len(sys.argv) > 1 else "gherkin"
    output_path = sys.argv[2] if len(sys.argv) > 2 else "escenarios.xlsx"

    if not os.path.isdir(gherkin_dir):
        print(f"Error: directorio '{gherkin_dir}' no encontrado")
        sys.exit(1)

    print(f"Escaneando: {gherkin_dir}")
    rows = collect_rows(gherkin_dir)
    print(f"Escenarios encontrados: {len(rows)}")
    write_excel(rows, output_path)


if __name__ == "__main__":
    main()
